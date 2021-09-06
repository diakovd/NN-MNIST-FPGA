# -*- coding: utf-8 -*-
"""
Created on Tue Mar 16 14:01:25 2021

@author: dyakov
"""
# ----------------------
# - read the input data:
import mnist_loader
training_data, validation_data, test_data = mnist_loader.load_data_wrapper()
training_data = list(training_data)

import gzip
import pickle
import numpy as np

with gzip.open('mnist.pkl.gz', 'rb') as f:
    train_set, valid_set, test_set = pickle.load(f, encoding='latin1')
    
train_x, train_y = train_set
data_in = train_x[2]

# ----------------------
# - Plot input data:
import matplotlib.cm as cm
import matplotlib.pyplot as plt
plt.imshow(data_in.reshape((28, 28)), cmap=cm.Greys_r)
plt.show()

data_in = np.reshape(data_in, (1, 784)) 
data_in = data_in.tolist()
#print(data_in)

# ----------------------
# - open exel book
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

wb = Workbook()

# grab the active worksheet
ws = wb.active

# ----------------------
#load network
import json
f = open("net_data.json", "r")
data = json.load(f)
f.close()

# ----------------------
#Write network to Exel Workbook
ws.append(["data_in"])
#print(data_in)
for i in data_in:
    ws.append(i)
    #print(ws._current_row)
    #print(get_column_letter(1))
ws.append(["weights"])

layer_dat = 2 #input data row for first laeyr

next_layer_dat = []
for w,b in zip(data["weights"],data["biases"]):
    for i,y in zip(w,b):
        ws.append(i)
        s = []
        for z in range(1,(len(i)+1)):
            s.append("=" + get_column_letter(z) + str(layer_dat) + '*' + get_column_letter(z) + str(ws._current_row))
        s.append("=SUM(" +  get_column_letter(1) + str(ws._current_row + 1) + ":" + get_column_letter(len(i))+ str(ws._current_row + 1) + ")")
        #data["biases"][0][0]
        s.append(str(y[0]))
        s.append("=1.0/(1.0+EXP(-("+ get_column_letter(len(i)+1) + str(ws._current_row + 1) + '+' + get_column_letter(len(i)+2) + str(ws._current_row + 1)+")))")
        next_layer_dat.append('='+ get_column_letter(len(i)+3) + str(ws._current_row + 1))
        ws.append(s)
        
        #1.0/(1.0+EXP(-z))
    #print(next_layer_dat) 
    ws.append(["next Laer"])
    ws.append(["Laer data"])
    layer_dat = ws._current_row + 1
    #print(layer_dat)
    ws.append(next_layer_dat)
    ws.append(["next Laer weights"])
    next_laer_dat = []

ws.append(["0","1","2","3","4","5","6","7","8","9"])
'''    
ws.append(["biases"])

print(data["biases"][0][0])

for b in data["biases"]:
    for i in b:
        ws.append(i)
'''

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])


# Save the file
wb.save("../output_files/net.xlsx")